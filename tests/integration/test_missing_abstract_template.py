"""GET /library/missing-abstracts-template: the "download, fill in
ABSTRACT, re-upload" round trip for whichever year the user is working on.

The header row must exactly match what dataset_import_service recognizes
-- these tests prove the whole loop closes, not just that a file downloads.
"""

import io

import openpyxl
import pytest
from sqlalchemy import select

from app.db.models import ReferencePaper
from app.db.models.enums import EnrichmentStatus
from tests.conftest import FIXTURES

pytestmark = pytest.mark.integration


async def _import_sample(client):
    data = (FIXTURES / "sample_dataset.xlsx").read_bytes()
    await client.post("/api/v1/library/import", files={"file": ("sample_dataset.xlsx", data)})


def _read_rows(xlsx_bytes: bytes) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


async def test_template_lists_only_rows_missing_an_abstract_for_that_year(client):
    await _import_sample(client)

    resp = await client.get("/api/v1/library/missing-abstracts-template?year=2021")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "missing-abstracts-2021.xlsx" in resp.headers["content-disposition"]

    rows = _read_rows(resp.content)
    header, *body = rows
    assert header == [
        "NO.",
        "YEAR",
        "FIELD OF STUDY",
        "AUTHORS",
        "DOCUMENT TITLE",
        "SOURCE TITLE",
        "LINK",
        "ABSTRACT",
    ]
    # Rows 4 and 9 are the two 2021 rows with no abstract in the fixture.
    assert {row[0] for row in body} == {4, 9}
    assert all(row[1] == 2021 for row in body)
    assert all(row[-1] is None for row in body)  # ABSTRACT left blank to fill in


async def test_template_for_unknown_year_omits_the_query_param(client, db_session):
    await _import_sample(client)
    ref = db_session.execute(
        select(ReferencePaper).where(ReferencePaper.original_row_number == 4)
    ).scalar_one()
    ref.year = None
    db_session.commit()

    resp = await client.get("/api/v1/library/missing-abstracts-template")
    assert resp.status_code == 200
    assert "missing-abstracts-unknown-year.xlsx" in resp.headers["content-disposition"]

    _, *body = _read_rows(resp.content)
    assert [row[0] for row in body] == [4]


async def test_filling_in_the_downloaded_template_and_reuploading_backfills(client, db_session):
    await _import_sample(client)

    resp = await client.get("/api/v1/library/missing-abstracts-template?year=2021")
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    filler = (
        "This filled-in abstract was pasted back from the downloaded template to "
        "demonstrate that the round trip actually backfills the missing field "
        "on the correct existing rows rather than creating duplicates."
    )
    for row in ws.iter_rows(min_row=2):
        row[-1].value = filler  # ABSTRACT is the last column

    buffer = io.BytesIO()
    wb.save(buffer)

    reupload = await client.post(
        "/api/v1/library/import",
        files={"file": ("filled_in.xlsx", buffer.getvalue())},
    )
    assert reupload.status_code == 201
    body = reupload.json()
    assert body["imported"] == 0
    assert body["backfilled_abstracts"] == 2

    refs = {
        r.original_row_number: r
        for r in db_session.execute(
            select(ReferencePaper).where(ReferencePaper.original_row_number.in_([4, 9]))
        ).scalars()
    }
    assert refs[4].abstract == filler
    assert refs[4].enrichment_status == EnrichmentStatus.ENRICHED
    assert refs[9].abstract == filler
